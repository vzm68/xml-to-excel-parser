import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# get import XML file
tree = ET.parse('template.xml')
root = tree.getroot()
offers = root.find('shop/offers')

# get import Excel file
wb = load_workbook(filename='offers.xlsx')
ws = wb.active

# update data from Excel
for row in ws.iter_rows(min_row=2, values_only=True):
    """Get data from row index"""
    offer_id = row[0]
    price = row[2]
    price_old = row[3]
    stock_quantity = row[4]
    # picture = row[6]
    for offer in offers.findall('offer'):
        """Loop through each element of the xml file"""
        if offer.find('vendorCode').text == offer_id:  # ID must be unique, I determine by it
            offer.find('price').text = str(price)
            offer.find('stock_quantity').text = str(stock_quantity)
            price_old_elem = offer.find('price_old')
            # offer.find('picture').text = picture
            if price_old_elem is None:
                """
                If price_old elem is absent and we don't need it. Leave as is.
                If the price and the old price are different, then it's time to make a change to the file :)
                """
                if str(price_old) != str(price):
                    price_old_elem = ET.Element('price_old')
                    price_old_elem.text = str(price_old)
                    offer.append(price_old_elem)
                    print(f"Added price_old={price_old} to offer with id={offer_id}")
            else:
                if str(price_old) == str(price):
                    offer.remove(price_old_elem)
                    print(f"Remove price_old={price_old} from offer with id={offer_id}")
                else:
                    price_old_elem.text = str(price_old)


# save data to export XML
tree.write('export.xml', encoding='utf-8', xml_declaration=True)